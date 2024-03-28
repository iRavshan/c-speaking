import openpyxl
import json
from django.shortcuts import render
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from user.decorators import user_access
from .models import Topic, Question, Answer

@login_required
@user_access
def Mock(request):
    questions = Question.objects.all()
    context = {
        'questions': questions
    }
    return render(request, 'speaking/mock.html', context)


@login_required
def Part1(request):
    context = {}
    if request.method == 'POST':
        topic_id = request.POST.get('topic_id')
        if int(topic_id) == -1:
           random_topic = Topic.objects.filter(question__part='1').distinct().order_by('?').first()
           topic_id = random_topic.id 
        random_questions = Question.objects.filter(topic_id=topic_id, part='1').order_by('?')[:6]
        if random_questions is not None:
            context['questions'] = random_questions
            return render(request, 'speaking/part1/part1.html', context)
    
    context['topics'] = Topic.objects.filter(question__part='1').distinct()
    return render(request, 'speaking/part1/topics.html', context)


@login_required
@user_access
def Part2(request):
    question = Question.objects.get(id=9)
    context = {
        'question': question,
    }
    return render(request, 'speaking/part2/part2.html', context)


@login_required
@user_access
def Part3(request):
    context = {}

    if request.method == 'POST':
        topic_id = request.POST.get('topic_id')
        if int(topic_id) == -1:
           random_topic = Topic.objects.filter(question__part='3').distinct().order_by('?').first()
           topic_id = random_topic.id
        random_questions = Question.objects.filter(topic_id=topic_id, part='3').order_by('?')[:6]
        if random_questions is not None:
            context['questions'] = random_questions
            return render(request, 'speaking/part3/part3.html', context)            
    
    context['topics'] = Topic.objects.filter(question__part='3').distinct()
    return render(request, 'speaking/part3/topics.html', context)


@login_required
def save_answers(request):
    if request.method == 'POST':

        part = request.POST.get('part')
        question_ids = request.POST.getlist('questions')
        answers = request.FILES.getlist('answers')

        answer = Answer(user=request.user, part=part)
        answer.save()

        for question_id in question_ids:
            question = Question.objects.get(id=question_id)
            answer.questions.add(question)
                    
        return render(request, 'core/home.html')
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=400)
    

def save_questions():
    path = "C://part1.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    data = wb_obj.active

    for col in range(1, 25):
        topic = Topic(name=(data.cell(row=1, column=col)).value)
        topic.save()
        for row in range(2, 11):
            cell_obj = data.cell(row=row, column=col)
            if cell_obj.value is not None:
                question = Question(title=cell_obj.value, topic=topic)
                question.save()
            else:
                break

